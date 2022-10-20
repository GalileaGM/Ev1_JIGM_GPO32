import csv
import datetime
from heapq import merge
import openpyxl
import os
import os.path
import random

idcheckcliente=[]
diccionarioClientes={}
#///
idchecksala=[]
tuplasala=()
diccionarioSalas={}
#///
idcheckreserv=[]
listafechas=[]
tuplareserv=()
diccionarioreserv={}
#///
turnos={1:"Matutino",2:"Vespertino",3:"Nocturno"}
listaEncontrados=[]
reservDisponibles=[]
#///
libro = openpyxl.Workbook()
hoja = libro["Sheet"]
hoja.title = "Reporte "
#///
diccionarioTest={}

def registrocliente():
    idclientes=max(diccionarioClientes.keys(), default=0)+1
    print('Clave del cliente es ', idclientes)#opcional
    cliente = input('Nombre del cliente: ')
    diccionarioTemp = {}
    diccionarioTemp = {idclientes:cliente}
    diccionarioClientes.update(diccionarioTemp)
    print("El cliente ha sido registrado")

def registrosala():
    idsala=max(diccionarioSalas.keys(), default=0)+1
    print('Clave de la sala es ', idsala)#opcional
    salas = (input('Nombre de la sala: ')) #crea nombre sala
    #x=0
    #while x==0:
    try:
        
        cupo_sala = int((input('Cupo de la sala: ')))
    except ValueError:
        print("Respuesta no valida: Solo se acepta numero")
        return
    tuplasala = (salas,cupo_sala)
    diccionarioTemp = {}
    diccionarioTemp = {idsala:tuplasala}
    diccionarioSalas.update(diccionarioTemp)
    print("La sala ha sido registrada correctamente")

def registroreservacion():
    if diccionarioSalas=={}:        #checa si hay salas registradas
            print('Error: No hay salas registradas')
            return
    if diccionarioClientes=={}:     #checa si hay clientes registrados
            print('Error: No hay clientes registrados')
            return
    fecha=[]
    try: #try para prevenir error de mal fecha ingresada
        fecha_capturada = input("Dime una fecha (dd/mm/aaaa): \n")
        fechareservacion = datetime.datetime.strptime(fecha_capturada, "%d/%m/%Y").date()   #fecha de reservacion
    except ValueError:
        print("Respuesta no valida: Solo se acepta fecha")
        return
    fecha.append(str(fechareservacion))                                                         #guarda fecha en lista temporal
    fechaactual=datetime.datetime.today().date()                                        #fecha actual
    fechalimite= fechareservacion + datetime.timedelta(days=-2)                         #fecha 2 dias antes de reservacion
    if fechaactual >= fechalimite:                                                      #Compara que la fecha sea valida
        print('Fecha no valida: Ingrese una fecha con dos dias de anticipacion.')
    else:
        idreserv=max(diccionarioreserv.keys(), default=0)+1
        nombrereservacion=(input('\nCual es el nombre de la reservacion: '))       #nombre de reservacion
        print('*'*14)                                                                   #menu para Turnos
        print('TURNOS')
        print('*'*14)
        print('[M]. Matutino')
        print('[V]. Vespertino')
        print('[N]. Nocturno' )
        opmenu=input('\nEscoge la opción: ')                                            #escoje opcion de menu para Turnos
        if (not opmenu.upper() in "MVN"):
            print("Opción no válida: Favor de seleccionar opcion en la lista.")
        if (opmenu.upper()=="M"):
            turno=1
            fecha.append(turnos[turno])
        elif (opmenu.upper()=="V"):
            turno=2
            fecha.append(turnos[turno])
        elif (opmenu.upper()=="N"):
            turno=3
            fecha.append(turnos[turno])
        for x in range(len(listafechas)):
            if str(listafechas[x])==str(fecha):
                print('Turno ya ocupado. No puede realizarse la reservacion.')           #checa si la fecha y turno son iguales
                return
        listafechas.append(fecha)
        print(diccionarioSalas)
        numerosala=input('Ingresa el numero de sala que quieres usar: ')
        print(diccionarioClientes)
        numerocliente=input('Ingresa tu id asignado: ')
        tuplareserv=(str(fecha[0]),int(numerosala),int(numerocliente),nombrereservacion,turno)  #dd/mm/aaaa,id sala,id cliente,nombre reserva,turno
        diccionarioTemp = {}
        diccionarioTemp = {idreserv:tuplareserv}
        diccionarioreserv.update(diccionarioTemp)
        print(diccionarioreserv)

def editarnombre():
    print(diccionarioreserv)
    idBuscado=int(input('Ingrese el id deseado: '))                                #pregunta por id a cambiar
    nuevoNombre=str(input('ingrese nombre nuevo: '))                               #pregunta por nuevo nombre
    templist = list(diccionarioreserv[idBuscado])
    templist[3]=nuevoNombre
    diccionarioreserv[idBuscado] = templist
    print(diccionarioreserv)

def consulta():
    try:
        fecha_capturada = input("Ingrese la fecha a consultar (dd/mm/aaaa): \n")
        fechareservacion = datetime.datetime.strptime(fecha_capturada, "%d/%m/%Y").date()
    except ValueError:
        print('Respuesta no valida: solo se acepta fecha')
        return
    listabusqueda = []                                                                                  #lista temporal para buscar en el diccionario
    templist = list(diccionarioreserv.keys())
    for x in templist:
        if diccionarioreserv[x][0] != str(fechareservacion):
            continue
        else:
            listabusqueda.append(x)
            continue
    print('-'*76)
    print('--\t\t\tREPORTE DE RESERVACIONES PARA EL DIA ', fechareservacion,'\t\t\t--')
    print('-'*76)
    print('--FECHA\t\t\t\tSALA\t\tCLIENTE\t\tEVENTO\t\tTURNO\t--')
    print('-'*76)
    if listabusqueda == []:
        print(' '*23, 'No hay eventos para ', fechareservacion)
    else:
        for x in listabusqueda:
            print(f'{diccionarioreserv[x][0]}\t\t\t{diccionarioSalas[diccionarioreserv[x][1]][0]}\t\t{diccionarioClientes[diccionarioreserv[x][2]]}\t\t{diccionarioreserv[x][3]}\t\t{diccionarioreserv[x][4]}') #diccionario reserva
    print('-'*76)
    print('--                                FIN DEL REPORTE                         --')

def exportarexcel():
    try:
        fecha_capturada = input("Ingrese la fecha a consultar (dd/mm/aaaa): \n")
        fechareservacion = datetime.datetime.strptime(fecha_capturada, "%d/%m/%Y").date()
    except ValueError:
        print('Respuesta no valida: solo se acepta fecha')
        return
    listabusqueda = []                                                                                  #lista temporal para buscar en el diccionario
    templist = list(diccionarioreserv.keys())
    for x in templist:
        if diccionarioreserv[x][0] != str(fechareservacion):
            continue
        else:
            listabusqueda.append(x)
        continue
    hoja["B1"].value = "Hoy es:"
    hoja["C1"].value = datetime.datetime.now().strftime("%d/%m/%Y")
    hoja.merge_cells('A2:K2')
    hoja["A2"].value=("--"*100)
    hoja.merge_cells('A3:B3')
    hoja["A3"].value=("--"*100)
    hoja.merge_cells('C3:G3')
    hoja["C3"].value="REPORTE DE RESERVACIONES PARA EL DIA "
    hoja.merge_cells('H3:I3')
    hoja["H3"].value=f"{fechareservacion}"
    hoja.merge_cells('J3:K3')
    hoja["J3"].value=("--"*100)
    hoja.merge_cells('A4:K4')
    hoja["A4"].value=("--"*100)
    hoja.merge_cells('A5:C5')
    hoja["A5"].value=("--"*100)
    hoja["D5"].value="FECHA"
    hoja["E5"].value="SALA"
    hoja["F5"].value="CLIENTE"
    hoja["G5"].value="EVENTO"
    hoja["H5"].value="TURNO"
    hoja.merge_cells('I5:K5')
    hoja["I5"].value=("--"*100)
    hoja.merge_cells('A6:K6')
    hoja["A6"].value=("--"*100)
    if listabusqueda ==[]:
        hoja.merge_cells('A7:C7')
        hoja["A7"].value=(" ")
        hoja.merge_cells('D7:F7')
        hoja["D7"].value="No hay eventos para"
        hoja.merge_cells('G7:H7')
        hoja["G7"].value=f"{fechareservacion}"
        hoja.merge_cells('I7:K7')
        hoja["I7"].value=("--"*100)
        
    else:
        for x in listabusqueda:
            hoja.merge_cells('A7:B7')
            hoja["A7"].value=("--"*100)
            hoja.merge_cells('C7:D7')
            hoja["C7"].value=f"{diccionarioreserv[x][0]}"
            hoja["E7"].value=f"{diccionarioSalas[diccionarioreserv[x][1]][0]}"
            hoja["F7"].value=f"{diccionarioClientes[diccionarioreserv[x][2]]}"
            hoja["G7"].value=f"{diccionarioreserv[x][3]}"
            hoja["H7"].value=f"{diccionarioreserv[x][4]}"
            hoja.merge_cells('I7:K7')
            hoja["I7"].value=("--"*100)
    hoja.merge_cells('A8:C8')
    hoja["A8"].value=("--"*100)
    hoja.merge_cells('D8:H8')
    hoja["D8"].value="FIN DEL REPORTE .... "
    hoja.merge_cells('I8:K8')
    hoja["I8"].value=("--"*100)

    libro.save(f"ReporteDeReservaciones{fechareservacion}.xlsx")
    print("\nLibro del reporte creado exitosamente!")

def consultasala():
    fechaInput = input("Dime una fecha (dd/mm/aaaa): \n")
    fechaBuscada = datetime.datetime.strptime(fechaInput, "%d/%m/%Y").date()
    for clave,valor in diccionarioreserv.items():   #crea lista de salas en uso
        sala,fecha,turno=(valor[1],valor[0],valor[4])
        if str(fecha) == str(fechaBuscada):
            listaEncontrados.append((sala,turno))
            reservEncontradas=set(listaEncontrados)
    for sala in diccionarioSalas.keys():            #crea lista de todas las posibilidades
        for turno in turnos.keys():
            reservDisponibles.append((sala,turno))
            combinacionReservDisponibles=set(reservDisponibles)
    salasTurnosDisponibles=sorted(combinacionReservDisponibles-reservEncontradas)
    print('*'*30)
    print(f"*Salas disponibles para rentar el {fechaBuscada.strftime('%d/%m/%Y')}*\n")
    print("Salas\t\tTurnos")
    for sala,turno in salasTurnosDisponibles:
        print(f"{sala},{diccionarioSalas[sala][0]}\t\t{turnos[turno]}")

#CARGA DE DATOS~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
#diccionarioClientes
if os.path.exists('diccionarioClientes.csv') == False:
    with open('diccionarioClientes.csv','w') as f:
        f.close()
with open('diccionarioClientes.csv','r',newline='') as archivo:
    reader = csv.reader(archivo)
    diccionarioTest.clear()
    diccionarioTest.update(reader)
    for llave in diccionarioTest.copy():
        diccionarioTest[int(llave)] = diccionarioTest[llave]
        del diccionarioTest[llave]
    diccionarioClientes.update(diccionarioTest)
    print(diccionarioClientes)

#diccionarioSalas
if os.path.exists('diccionarioSalas.csv') == False:
    with open('diccionarioSalas.csv','w') as f:
        f.close()
with open('diccionarioSalas.csv','r',newline='') as archivo:
    reader = csv.reader(archivo)
    diccionarioTest.clear()
    diccionarioTest.update(reader)
    for llave in diccionarioTest.copy():
        diccionarioTest[int(llave)] = diccionarioTest[llave]
        del diccionarioTest[llave]
        diccionarioTest[int(llave)] = eval(diccionarioTest[int(llave)])
    diccionarioSalas.update(diccionarioTest)
    print(diccionarioSalas)

#diccionarioreserv
if os.path.exists('diccionarioreserv.csv') == False:
    with open('diccionarioreserv.csv','w') as f:
        f.close()
with open('diccionarioreserv.csv','r',newline='') as archivo:
    reader = csv.reader(archivo)
    diccionarioTest.clear()
    diccionarioTest.update(reader)
    for llave in diccionarioTest.copy():
        diccionarioTest[int(llave)] = diccionarioTest[llave]
        del diccionarioTest[llave]
        diccionarioTest[int(llave)] = eval(diccionarioTest[int(llave)])
    diccionarioreserv.update(diccionarioTest)
    print(diccionarioreserv)

#listafechas
if os.path.exists('listafechas.csv') == False:
    with open('listafechas.csv','w') as f:
        f.close()
with open('listafechas.csv','r',newline='') as archivo:
    reader = csv.reader(archivo)
    listafechas = list(reader)
    print(listafechas)
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

while True:
    print('*'*14)
    print('MENU DE OPCIONES')
    print('*'*14)
    print('[A]. Registrar cliente')
    print('[B]. Registrar sala ')
    print('[C]. Reservación' )
    print('[D]. Reporte de reservaciones')
    print('[E]. SALIR')

    opmenu=input('\nEscoge la opción: ')

    if (not opmenu.upper() in "ABCDE"):
        print("Opción no válida, intenta de nuevo.")
        continue

    #OPCION=SALIR
    if (opmenu.upper()=="E"):
        
        with open('diccionarioSalas.csv','w',newline='') as archivo:    #Guarda diccionarioSalas
            writer = csv.writer(archivo)
            for key, value in diccionarioSalas.items():
                writer.writerow([key, value])
        
        with open('diccionarioClientes.csv','w',newline='') as archivo: #Guarda diccionarioClientes
            writer = csv.writer(archivo)
            for key, value in diccionarioClientes.items():
                writer.writerow([key, value])
        
        with open('diccionarioreserv.csv','w',newline='') as archivo: #Guarda diccionarioreserv
            writer = csv.writer(archivo)
            for key, value in diccionarioreserv.items():
                writer.writerow([key, value])
        
        with open('listafechas.csv','w',newline='') as archivo:
            writer = csv.writer(archivo)
            for value in listafechas:
                writer.writerow(value)
        
        print("Fin de la ejecución.")
        break

    #OPCION=REGISTRAR CLIENTE
    if (opmenu.upper()=="A"):
        registrocliente()

    #OPCION=REGISTRAR SALA
    if (opmenu.upper()=="B"):
        registrosala()

    #OPCION=SUBMENU RESERVACION
    if (opmenu.upper()=="C"):
        while True:
            print('*'*14)
            print('SUBMENU DE RESERVACIONES')
            print('*'*14)
            print('A. REGISTRAR NUEVA RESERVACIÓN')
            print('B. MODIFICAR NOMBRE RESERVACIÓN')
            print('C. CONSULTAR DISPONIBILIDAD DE SALA')
            print('D. REGRESAR AL MENU')
            submenur=input('\nEscoge la opción: ')

            if (not submenur.upper() in "ABCD"):
                print("Opción no válida, intenta de nuevo.")
                continue
            if (submenur.upper()=="A"):
                registroreservacion()

            if (submenur.upper()=="B"):
                editarnombre()

            if (submenur.upper()=="C"):
                consultasala()

            if (submenur.upper()=="D"):
                break
                continue

    #OPCION=SUBMENU REPORTE
    if (opmenu.upper()=="D"):
        while True:
            print('*'*14)
            print('SUBMENU DE REPORTE')
            print('*'*14)
            print('A. REPORTE EN PANTALLA PARA RESERVACION DE UNA FECHA')
            print('B. EXPORTAR REPORTE EN EXCEL')
            print('C. REGRESAR AL MENU')
            submenure=input('\nEscoge la opción: ')

            if (not submenure.upper() in "ABCD"):
                print("Opción no válida, intenta de nuevo.")
                continue
            if (submenure.upper()=="A"):
                consulta()
            if (submenure.upper()=="B"):
                exportarexcel()
            if (submenure.upper()=="C"):
                break
                continue
