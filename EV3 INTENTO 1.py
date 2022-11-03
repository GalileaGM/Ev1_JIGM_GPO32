import csv
import datetime
from heapq import merge
import openpyxl
import os
import os.path
import random
import sqlite3
from sqlite3 import Error
import sys

encontradas = []
disponibles = []

#///
libro = openpyxl.Workbook()
hoja = libro["Sheet"]
hoja.title = "Reporte "
#/// CREACION DE BASE DE DATOS Y TABLAS
if not os.path.isfile("EVIDENCIA3.db"):
    print("Ejecucion Inicial")
    try:
        with sqlite3.connect("EVIDENCIA3.db") as conn:
            cursor=conn.cursor()
            cursor.execute("CREATE TABLE IF NOT EXISTS cliente (Clavecliente INTEGER PRIMARY KEY, Nombrecliente TEXT NOT NULL);")
            cursor.execute("CREATE TABLE IF NOT EXISTS sala (Clavesala INTEGER PRIMARY KEY, Nombresala TEXT NOT NULL, Cuposala INTEGER NOT NULL);")
            cursor.execute("CREATE TABLE IF NOT EXISTS turno (Claveturno INTEGER PRIMARY KEY, Nombreturno TEXT NOT NULL);")
            cursor.execute("CREATE TABLE IF NOT EXISTS reservacion (Clavereservacion INTEGER PRIMARY KEY, Nombrecliente TEXT NOT NULL, Nombrereservacion TEXT NOT NULL, Fechareservacion timestamp, turnore INTEGER NOT NULL, salare INTEGER NOT NULL,FOREIGN KEY(turnore) REFERENCES turno(Claveturno), FOREIGN KEY(salare) REFERENCES sala(Clavesala));")
            print("Tablas creadas")
            cursor.execute("INSERT INTO turno VALUES(1,'matutino')")
            cursor.execute("INSERT INTO turno VALUES(2,'vespertino')")
            cursor.execute("INSERT INTO turno VALUES(3,'nocturno')")
            print("Turnos creados")
    except Error as e:
        print(e)
    except:
        print(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
    finally:
        conn.close()
if os.path.isfile("EVIDENCIA3.db"):
    print("Base de datos existente")
    while True:
        print('*'*14)
        print('MENU DE OPCIONES')
        print('*'*14)
        print('[A]. Registrar cliente')
        print('[B]. Registrar sala ')
        print('[C]. Reservación' )
        print('[D]. Reporte de reservaciones')
        print('[E]. SALIR')

        opmenu=input('\nEscoge la opción: ')

        if (not opmenu.upper() in "ABCDE"):
            print("Opción no válida, intenta de nuevo.")
            continue

        #OPCION=SALIR
        if (opmenu.upper()=="E"):
            print("Fin de la ejecución.")
            break

        #OPCION=REGISTRAR CLIENTE
        if (opmenu.upper()=="A"):
            while True:
                Nombrecliente=str(input("Ingrese el nombre del cliente : "))
                if Nombrecliente=='':
                    print("No debe omitirse el nombre del cliente")
                    continue
                else:
                    try:
                        with sqlite3.connect("EVIDENCIA3.db") as conn:
                            cursor=conn.cursor()
                            valores={"Nombrecliente":Nombrecliente}
                            cursor.execute("INSERT INTO cliente (Nombrecliente) VALUES(:Nombrecliente)", valores)
                            print("Registro agregado exitosamente")
                    except Error as e:
                        print(e)
                    except:
                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                    finally:
                        conn.close()
                break
            continue

        #OPCION=REGISTRAR SALA
        if (opmenu.upper()=="B"):
            while True:
                Nombresala=str(input("Ingrese el nombre de la sala : "))
                if Nombresala=='':
                    print("No debe omitirse el nombre de la sala")
                    continue
                else:
                    while True:
                        try:
                            Cuposala=int(input("Ingrese el cupo de la sala: "))
                        except ValueError:
                            print("Respuesta no valida: Solo se acepta numero")
                            continue
                        if Cuposala<=0:
                            print("Incorrecto, el cupo de la sala debe de ser mayor a 0")
                            continue
                        else:
                            try:
                                with sqlite3.connect("EVIDENCIA3.db") as conn:
                                    cursor=conn.cursor()
                                    valores={"Nombresala":Nombresala, "Cuposala":Cuposala}
                                    cursor.execute("INSERT INTO sala (Nombresala, Cuposala) VALUES(:Nombresala,:Cuposala)", valores)
                                    print("Registro ha sido exitoso")
                            except Error as e:
                                print(e)
                            except:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                            finally:
                                conn.close()
                        break
                break
            continue

        #OPCION=SUBMENU RESERVACION
        if (opmenu.upper()=="C"):
            while True:
                print('*'*14)
                print('SUBMENU DE RESERVACIONES')
                print('*'*14)
                print('A. REGISTRAR NUEVA RESERVACIÓN')
                print('B. MODIFICAR NOMBRE RESERVACIÓN')
                print('C. CONSULTAR DISPONIBILIDAD DE SALA')
                print('D. ELIMINAR RESERVACIÓN')
                print('F. REGRESAR AL MENU')
                submenur=input('\nEscoge la opción: ')

                if (not submenur.upper() in "ABCDF"):
                    print("Opción no válida, intenta de nuevo.")
                    continue
                if (submenur.upper()=="A"):
                    while True:
                        try:
                            with sqlite3.connect("EVIDENCIA3.db") as conn:
                                cursor=conn.cursor()
                                cursor.execute("SELECT * FROM cliente ORDER BY Clavecliente")
                                regcliente = cursor.fetchall()
                                #Procedemos a evaluar si hay registros en la respuesta
                                if regcliente:
                                    print("\tCLIENTES REGISTRADOS")
                                    print("Clavecliente\tNombrecliente")
                                    print("--" * 50)
                                    for Clavecliente, Nombrecliente in regcliente:
                                        print(f"{Clavecliente:^10} \t {Nombrecliente}")

                                    Nomcliente=str(input("Ingrese el nombre del cliente a buscar : "))
                                    if Nomcliente=='':
                                        print("No debe omitirse el nombre del cliente")
                                        continue
                                    else:
                                        valores = {"Nombrecliente":Nomcliente}
                                        cursor.execute("SELECT * FROM cliente WHERE Nombrecliente = :Nombrecliente", valores)
                                        regcliente = cursor.fetchall()
                                        if regcliente:
                                            for Clavecliente, Nomcliente in regcliente:
                                                print("Se encuentra registrado")
                                                continue
                                        else:
                                            print(f"No esta registrado, favor de registrarse")
                                            break
                                        #SI SE ENCUENTRA REGISTRADO SE COMIENZA EL REGISTRO RESERVACION
                                        while True:
                                            fecha_actual = datetime.date.today()
                                            Fechareservacion = fecha_actual
                                            while True:
                                                Nombrereservacion=str(input("Ingrese el nombre de la reservacion : "))
                                                if Nombrereservacion=='':
                                                    print("No debe omitirse el nombre de la reservacion")
                                                    continue
                                                else:
                                                    while True:
                                                        try:
                                                            Fechareservacion= input("Ingrese la fecha de reservacion : ")
                                                            if Fechareservacion=='':
                                                                print("No debe omitirse la fecha de la reservacion")
                                                                continue
                                                            else:
                                                                Fechareserva=datetime.datetime.strptime(Fechareservacion, "%d/%m/%Y" ).date()
                                                        except ValueError:
                                                                print("Respuesta no valida: no se cumple con el formato")
                                                                continue
                                                        fechalimite= Fechareserva + datetime.timedelta(days=-2)
                                                        if fecha_actual >= fechalimite:
                                                            print('Fecha no valida: Ingrese una fecha con dos dias de anticipacion.')
                                                            continue
                                                        #se realizan los insert
                                                        else:
                                                            while True:
                                                                try:
                                                                    with sqlite3.connect("EVIDENCIA3.db") as conn:
                                                                        cursor = conn.cursor()
                                                                        cursor.execute("SELECT * FROM turno ORDER BY Claveturno")
                                                                        regturno = cursor.fetchall()
                                                                        #Procedemos a evaluar si hay registros en la respuesta
                                                                        if regturno:
                                                                            print("\nTURNOS....")
                                                                            print("Claveturno\tNombreturno")
                                                                            print("--" * 30)
                                                                            for Claveturno, Nombreturno in regturno:
                                                                                print(f"{Claveturno:^10}\t{Nombreturno}")
                                                                            turnore=int(input("Ingrese el turno que desea reservar : "))
                                                                            valores = {"Claveturno":turnore}
                                                                            cursor.execute("SELECT * FROM turno WHERE Claveturno = :Claveturno", valores)
                                                                            regturno = cursor.fetchall()
                                                                            if regturno:
                                                                                for Claveturno, turnore in regturno:
                                                                                    print("SE REGISTRO CORRECTAMENTE")
                                                                                    continue
                                                                                while True:
                                                                                    try:
                                                                                        with sqlite3.connect("EVIDENCIA3.db") as conn:
                                                                                            cursor = conn.cursor()
                                                                                            cursor.execute("SELECT * FROM sala ORDER BY Clavesala")
                                                                                            regsala = cursor.fetchall()
                                                                                            if regsala:
                                                                                                print("\tSALAS REGISTRADAS....")
                                                                                                print("Clavesala\tNombresala\tCuposala")
                                                                                                print("--" * 30)
                                                                                                for Clavesala, Nombresala, Cuposala in regsala:
                                                                                                    print(f"{Clavesala:^10}\t{Nombresala:^10}\t{Cuposala}")

                                                                                                salare=int(input("Ingresa el numero de la sala que desea reservar : "))
                                                                                                valores = {"Clavesala":salare}
                                                                                                cursor.execute("SELECT * FROM sala WHERE Clavesala = :Clavesala", valores)
                                                                                                regsala = cursor.fetchall()
                                                                                                if regsala:
                                                                                                    for Clavesala, salare, Cuposala in regsala:
                                                                                                        print("SE REGISTRO CORRECTAMENTE")
                                                                                                        continue
                                                                                                    try:
                                                                                                        with sqlite3.connect("EVIDENCIA3.db") as conn:
                                                                                                            cursor=conn.cursor()
                                                                                                            valores={"Nombrecliente":Nomcliente,"Nombrereservacion":Nombrereservacion, "Fechareservacion":Fechareservacion, "turnore":turnore, "salare":salare}
                                                                                                            cursor.execute("INSERT INTO reservacion (Nombrecliente,Nombrereservacion, Fechareservacion,turnore,salare) VALUES(:Nombrecliente,:Nombrereservacion,:Fechareservacion,:turnore,:salare)", valores)
                                                                                                            print("Registro ha sido exitoso")
                                                                                                    except Error as e:
                                                                                                        print(e)
                                                                                                    except:
                                                                                                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                                                                                    finally:
                                                                                                        conn.close()
                                                                                                        break
                                                                                                else:
                                                                                                    print("La opcion no es valida")
                                                                                                    continue
                                                                                            else:
                                                                                                print("No se encontraron registros en la respuesta")
                                                                                                continue
                                                                                    except Error as e:
                                                                                        print(e)
                                                                                    except Exception:
                                                                                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                                                                    finally:
                                                                                        conn.close()
                                                                                        break
                                                                                    
                                                                            else:
                                                                                print("Revisar su opcion de turno seleccionado")
                                                                                continue
                                                                        else:
                                                                            print("No se encontraron registros en la respuesta")
                                                                except Error as e:
                                                                    print(e)
                                                                except Exception:
                                                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                                                finally:
                                                                    conn.close()
                                                                break
                                                            break
                                                    break
                                            break
                                break
                        except Error as e:
                            print(e)
                        except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                        finally:
                            conn.close()
                            break
                        break
                    continue

                if (submenur.upper()=="B"):
                    while True:
                        try:
                            with sqlite3.connect("EVIDENCIA3.db") as conn:
                                cursor = conn.cursor()
                                cursor.execute("SELECT * FROM reservacion ORDER BY Clavereservacion")
                                regreserva = cursor.fetchall()
                                if regreserva:
                                    print("")
                                    print("{:<20} {:<20} {:<20} {:<40} {:<20} {:<20}".format('Clavereservacion','Nombrecliente','Nombrereservacion','Fechareservacion','turnore','salare' ))
                                    print("*" * 150)
                                    for Clavereservacion,Nombrecliente,Nombrereservacion,Fechareservacion,turnore,salare in regreserva:
                                        print("{:<20} {:<20} {:<20} {:<40} {:<20} {:<20}".format(Clavereservacion, Nombrecliente,Nombrereservacion,Fechareserva,turnore,salare ))
                                        continue
                                buscarclave=int(input("Ingrese la clave del registro que va a editar : "))
                                valor={"Clavereservacion":buscarclave}
                                cursor.execute("SELECT * FROM reservacion WHERE Clavereservacion = :Clavereservacion",valor)
                                regreserva = cursor.fetchall()
                                if regreserva:
                                    print("")
                                    nuevo_nombre=input("Ingrese el nuevo nombre de la reservacion  : ")
                                    valores = {"Nombrereservacion":nuevo_nombre,"Clavereservacion":buscarclave}
                                    cursor.execute("UPDATE reservacion SET Nombrereservacion = :Nombrereservacion WHERE Clavereservacion = :Clavereservacion",valores)
                                    regreserva = cursor.fetchall()
                                    print("El registro se edito exitosamente")
                                    break
                                else:
                                    if not Clavereservacion in regreserva:
                                        print("No existe este registro")
                                        continue
                        except Error as e:
                            print (e)
                        except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                        finally:
                            conn.close()
                        break

                if (submenur.upper()=="C"):
                    while True:
                        #try:
                            fechadisponible = input("Ingrese la fecha del reservacion (dd/mm/aaaa): ")
                            if fechadisponible=='':
                                print("No debe omitirse la fecha de la reservacion")
                                continue
                            else:
                                fechadisponible = datetime.datetime.strptime(fechadisponible, "%d/%m/%Y")
                                #try:
                                with sqlite3.connect("EVIDENCIA3.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                                    cursor = conn.cursor()
                                    valor = {"fecha":fechadisponible}
                                    cursor.execute("SELECT * FROM reservacion WHERE DATE(Fechareservacion) = :fecha;", valor)
                                    regreserva = cursor.fetchall()
                                    cursor.execute("SELECT * FROM turno")
                                    regturno = cursor.fetchall()
                                    cursor.execute("SELECT * FROM sala")
                                    regsala = cursor.fetchall()
                                    if regreserva:
                                        for Clavereservacion,Nombrecliente,Nombrereservacion,Fechareserva,turnore,salare in regreserva:
                                            Fechareserva,turnore,salare = (Fechareserva,turnore,salare)
                                            encontradas.append((salare,turnore))
                                            reservaciones_ocupadas = set(encontradas)
                                        for Clavesala,Nombresala,Cuposala in regsala:
                                            for Claveturno,Nombreturno in regturno:
                                                disponibles.append((salare,turnore))
                                                combinaciones_reservaciones_disponibles = set(disponibles)
                                        salas_turnos_disponibles = sorted(list(combinaciones_reservaciones_disponibles - reservaciones_ocupadas))
                                        print("\n Las opciones disponibles para rentar en esa fecha son : ")
                                        print(f"Salas que se encuentran disponibles para rentar el {fechadisponible}\n")
                                        print("Salas\t\tTurnos")
                                        if regsala:
                                            for salare,turnore in salas_turnos_disponibles:
                                                print(f"{salare}\t\t{turnore}")
                                #except Error as e:
                                    #print(e)
                                #except:
                                    #print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                #finally:
                                    #conn.close()
                        #except ValueError:
                            #print("Respuesta no valida: no se cumple con el formato")
                            #continuec
                            break
                    continue

                if (submenur.upper()=="D"):
                    while True:
                        try:
                            with sqlite3.connect("EVIDENCIA3.db") as conn:
                                cursor = conn.cursor()
                                cursor.execute("SELECT * FROM reservacion ORDER BY Clavereservacion")
                                regreserva = cursor.fetchall()
                                if regreserva:
                                    print("")
                                    print("{:<20} {:<20} {:<20} {:<40} {:<20} {:<20}".format('Clavereservacion','Nombrecliente','Nombrereservacion','Fechareservacion','turnore','salare' ))
                                    print("*" * 100)
                                    for Clavereservacion, Nombrecliente,Nombrereservacion,Fechareservacion,turnore,salare in regreserva:
                                        print("{:<20} {:<20} {:<20} {:<40} {:<20} {:<20}".format(Clavereservacion, Nombrecliente,Nombrereservacion,Fechareservacion,turnore,salare ))

                                eliminarregistro=int(input("Ingrese la clave del registro que desea eliminar : "))
                                with sqlite3.connect("EVIDENCIA3.db") as conn:
                                    cursor = conn.cursor()
                                    valor = {"Clavereservacion":eliminarregistro}
                                    cursor.execute("SELECT * FROM RESERVACION WHERE Clavereservacion = :Clavereservacion",valor)
                                    regreserva = cursor.fetchall()
                                    if regreserva:
                                        print("")
                                        print("{:<20} {:<20} {:<20} {:<40} {:<20} {:<20}".format('Clavereservacion','Nombrecliente','Nombrereservacion','Fechareservacion','turnore','salare' ))
                                        print("*" * 100)
                                        for Clavereservacion, Nombrecliente,Nombrereservacion,Fechareservacion,turnore,salare in regreserva:
                                            print("{:<20} {:<20} {:<20} {:<40} {:<20} {:<20}".format(Clavereservacion, Nombrecliente,Nombrereservacion,Fechareservacion,turnore,salare ))

                                    else:
                                        if not Clavereservacion in regreserva:
                                            print("No existe esta clave")
                                            continue
                                    print("")
                                    confirmar =input("Desea eliminar la reservacion [Si/No ]: ")
                                    if (not confirmar.upper() in "SN"):
                                        print("Opcion no valida")
                                    if (confirmar.upper() == "S"):
                                        cursor.execute("DELETE FROM reservacion WHERE Clavereservacion = :Clavereservacion",valor)
                                        print("Se ha eliminado la reservacion ")
                                        break
                                    elif (confirmar.upper() == "N"):
                                        print("")
                                        print("1. Desea regresar al menu principal")
                                        print("2. Eliminar otro registro")
                                        print("")
                                        regresar = int(input("Ingrese su opcion : "))
                                        if regresar == 1:
                                            break
                                        else:
                                            continue
                        except Error as e:
                            print(e)
                        except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                        finally:
                            conn.close()
                    continue

                if (submenur.upper()=="F"):
                    break
                continue

        #OPCION=SUBMENU REPORTE
        if (opmenu.upper()=="D"):
            while True:
                print('*'*14)
                print('SUBMENU DE REPORTE')
                print('*'*14)
                print('A. REPORTE EN PANTALLA PARA RESERVACION DE UNA FECHA')
                print('B. EXPORTAR REPORTE EN EXCEL')
                print('C. REGRESAR AL MENU')
                submenure=input('\nEscoge la opción: ')

                if (not submenure.upper() in "ABC"):
                    print("Opción no válida, intenta de nuevo.")
                    continue

                if (submenure.upper()=="A"):
                    while True:
                        fechareporte = input("Ingrese la fecha del evento (dd/mm/aaaa): ")
                        fechareporte = datetime.datetime.strptime(fechareporte, "%d/%m/%Y").date()
                        try:
                            with sqlite3.connect("EVIDENCIA3.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                                cursor = conn.cursor()
                                valor = {"fechareserva":fechareporte}
                                cursor.execute("SELECT Nombrecliente,Nombrereservacion,Fechareservacion,turnore,salare FROM reservacion WHERE DATE(Fechareservacion) = :fechareserva;", valor)
                                regreserva = cursor.fetchall()
                                print("--"*50)
                                print("--" + " "*30 + f"REPORTE DE RESERVACIONES PARA EL DÍA {fechareporte}" + " "*30 + "--")
                                print("--"*50)
                                print("{:<30} {:<30}  {:<30} {:<30}".format('Sala','Nombre Cliente','Nombre Reservacion', 'Turno' ))
                                print("--"*50)
                                if regreserva:
                                    for Nombrecliente,Nombrereservacion,Fechareservacion,turnore,salare in regreserva:
                                        print(f"{salare}\t\t\t{Nombrecliente}\t\t\t{Nombrereservacion}\t\t\t{turnore}")   
                                    print("")
                                    print("--"*50 + " FIN DEL REPORTE " + "--"*50)
                                else:
                                    print("No existen reservas con esa fecha")
                        except sqlite3.Error as e:
                            print (e)
                        except Exception:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                        finally:
                            if (conn):
                                conn.close()
                            continue
                if (submenure.upper()=="B"):
                    while True:
                        fechaexcel = input("Ingrese la fecha del evento (dd/mm/aaaa): ")
                        fechaexcel = datetime.datetime.strptime(fechaexcel, "%d/%m/%Y")
                        try:
                            with sqlite3.connect("EVIDENCIA3.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                                cursor = conn.cursor()
                                valor = {"fecha":fechaexcel}
                                cursor.execute("SELECT Nombrecliente,Nombrereservacion,Fechareservacion,turnore,salare FROM RESERVACION WHERE DATE(Fechareservacion) = :fechaexcel;", valor)
                                regreserva = cursor.fetchall()

                            if regreserva:
                                elementos_sala = [(salare,Nombrecliente,Nombrereservacion,turnore)]
                                hoja["B1"].value = "Hoy es:"
                                hoja["C1"].value = datetime.datetime.now().strftime("%d/%m/%Y")
                                hoja.merge_cells('A2:K2')
                                hoja["A2"].value=("--"*100)
                                hoja.merge_cells('A3:B3')
                                hoja["A3"].value=("--"*100)
                                hoja.merge_cells('C3:G3')
                                hoja["C3"].value="REPORTE DE RESERVACIONES PARA EL DIA "
                                hoja.merge_cells('H3:I3')
                                hoja["H3"].value=f"{fechaexcel}"
                                hoja.merge_cells('J3:K3')
                                hoja["J3"].value=("--"*100)
                                hoja.merge_cells('A4:K4')
                                hoja["A4"].value=("--"*100)
                                hoja.merge_cells('A5:C5')
                                hoja["A5"].value=("--"*100)
                                hoja["D5"].value="FECHA"
                                hoja["E5"].value="SALA"
                                hoja["F5"].value="CLIENTE"
                                hoja["G5"].value="EVENTO"
                                hoja["H5"].value="TURNO"
                                hoja.merge_cells('I5:K5')
                                hoja["I5"].value=("--"*100)
                                hoja.merge_cells('A6:K6')
                                hoja["A6"].value=("--"*100)
                                if fechaexcel ==[]:
                                    hoja.merge_cells('A7:C7')
                                    hoja["A7"].value=(" ")
                                    hoja.merge_cells('D7:F7')
                                    hoja["D7"].value="No hay eventos para"
                                    hoja.merge_cells('G7:H7')
                                    hoja["G7"].value=f"{fechaexcel}"
                                    hoja.merge_cells('I7:K7')
                                    hoja["I7"].value=("--"*100)
                                else:
                                    for Nombrecliente,Nombrereservacion,Fechareserva,turnore,salare in regreserva:
                                        sala = [(salare,Nombrecliente,Nombrereservacion,turnore)]
                                        for elemento in sala:
                                            hoja.append(elemento)
                                        libro.save("Reporte.xlsx")
                                        print("Libro creado exitosamente")
                        except sqlite3.Error as e:
                            print (e)
                        except Exception:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                        finally:
                            if (conn):
                                conn.close()
                            continue
                if (submenure.upper()=="C"):
                    break
                continue
