import csv
import datetime
from heapq import merge
import openpyxl
import os
import os.path
import random
import sqlite3
from sqlite3 import Error
import sys
#///
libro = openpyxl.Workbook()
hoja = libro["Sheet"]
hoja.title = "Reporte "
listaEncontrados=[]
reservDisponibles=[]
if not os.path.isfile("PIA.db"):
    print("Ejecucion Inicial")
    try:
        with sqlite3.connect("PIA.db") as conn:
            cursor=conn.cursor()
            cursor.execute("CREATE TABLE IF NOT EXISTS registrocliente (idcliente INTEGER PRIMARY KEY, clientes TEXT NOT NULL);")
            cursor.execute("CREATE TABLE IF NOT EXISTS registrosala (idsala INTEGER PRIMARY KEY, salas TEXT NOT NULL, cupo_sala INTEGER NOT NULL);")
            cursor.execute("CREATE TABLE IF NOT EXISTS TURNO (idturno INTEGER PRIMARY KEY, Nombreturno TEXT NOT NULL);")
            cursor.execute("CREATE TABLE IF NOT EXISTS RESERVACION(idreserv INTEGER PRIMARY KEY, nombrereservacion TEXT NOT NULL,fechareservacion timestamp,idcliente INTEGER NOT NULL, idturno INTEGER NOT NULL,idsala INTEGER NOT NULL,FOREIGN KEY(idcliente)REFERENCES registrocliente(idcliente),FOREIGN KEY(idturno) REFERENCES TURNO(idturno), FOREIGN KEY(idsala) REFERENCES registrosala(idsala));")
            #print("Tablas creadas")
            cursor.execute("INSERT INTO TURNO VALUES(1,'matutino')")
            cursor.execute("INSERT INTO TURNO VALUES(2,'vespertino')")
            cursor.execute("INSERT INTO TURNO VALUES(3,'nocturno')")
            #print("Turnos creados")
    except Error as e:
        print(e)
    except:
        print(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
    finally:
        conn.close()
if os.path.isfile("PIA.db"):
  while True:
    print("\n")
    print('*'*14)
    print('MENU DE OPCIONES')
    print('*'*14)
    print('[A]. Registrar cliente')
    print('[B]. Registrar sala ')
    print('[C]. Reservación' )
    print('[D]. Reporte de reservaciones')
    print('[E]. SALIR')
    opmenu=input('\nEscoge la opción: ')
    if (not opmenu.upper() in "ABCDE"):
      print("Opción no válida, intenta de nuevo.")
      continue
    #OPCION=SALIR
    if (opmenu.upper()=="E"):
      print("Fin de la ejecución.")
      break
    if (opmenu.upper()=="A"):
      while True:
        Nomcliente=str(input("Ingrese el nombre del cliente : "))
        if Nomcliente=='':
          print("No debe omitirse el nombre del cliente")
          continue
        else:
          try:
            with sqlite3.connect("PIA.db") as conn:
              cursor=conn.cursor()
              valores={"Nombrecliente":Nomcliente}
              cursor.execute("INSERT INTO registrocliente (clientes) VALUES(:Nombrecliente)", valores)
              print("Su registro ha sido exitosamente")
          except Error as e:
            print(e)
          except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
          finally:
            conn.close()
          break
      continue
    #OPCION=REGISTRAR SALA
    if (opmenu.upper()=="B"):
      while True:
        Nomsala=str(input("Ingrese el nombre de la sala : "))
        if Nomsala=='':
          print("No debe omitirse el nombre de la sala")
          continue
        else:
          while True:
            try:
              CupSala=int(input("Ingrese el cupo de la sala: "))
            except ValueError:
              print("Respuesta no valida: Solo se acepta numero")
              continue
            if CupSala<=0:
              print("Incorrecto, el cupo de la sala debe de ser mayor a 0")
              continue
            else:
              try:
                with sqlite3.connect("PIA.db") as conn:
                  cursor=conn.cursor()
                  valores={"Nombresala":Nomsala, "Cuposala":CupSala}
                  cursor.execute("INSERT INTO  registrosala(salas,cupo_sala) VALUES(:Nombresala,:Cuposala)", valores)
                  print("Registro ha sido exitoso")
              except Error as e:
                print(e)
              except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
              finally:
                conn.close()
              break
        break
      continue
    #REGISTRO RESERVACION
    #OPCION=SUBMENU RESERVACION
    if (opmenu.upper()=="C"):
      while True:
        print('*'*14)
        print('SUBMENU DE RESERVACIONES')
        print('*'*14)
        print('A. REGISTRAR NUEVA RESERVACIÓN')
        print('B. MODIFICAR NOMBRE RESERVACIÓN')
        print('C. CONSULTAR DISPONIBILIDAD DE SALA')
        print('D. ELIMINAR RESERVACIÓN')
        print('E. REGRESAR AL MENU')
        submenur=input('\nEscoge la opción: ')

        if (not submenur.upper() in "ABCDE"):
          print("Opción no válida, intenta de nuevo.")
          continue
        if (submenur.upper()=="F"):
          break
          continue
        if (submenur.upper()=="C"):
          while True:
            try:
              fechaInput = input("Dime una fecha (dd/mm/aaaa): \n")
              if fechaInput =='':
                print("No debe omitirse la fecha de la reservacion")
                continue
              else:
                fechaBuscada = datetime.datetime.strptime(fechaInput, "%d/%m/%Y").date()
                try:
                  with sqlite3.connect("PIA.db") as conn:
                    cursor = conn.cursor()
                    valor = {"fecha":fechaBuscada}
                    cursor.execute("SELECT * FROM RESERVACION WHERE DATE(fechareservacion) = :fecha;", valor)
                    regreserva = cursor.fetchall()
                    cursor.execute("SELECT idturno FROM TURNO")
                    regturno = cursor.fetchall()
                    cursor.execute("SELECT idsala FROM registrosala")
                    regsala = cursor.fetchall()
                    if regreserva:
                      if regturno:
                        if regsala:
                          for idreserv ,nombrereservacion,fechareservacion,idcliente,idturno,idsala in regreserva:
                              fechareservacion,idturno,idsala = (fechareservacion,idturno,idsala)
                              listaEncontrados.append((idsala,idturno))
                              reservaciones_ocupadas = set(listaEncontrados)
                          for idsala,salas,cupo_sala in regsala:
                              for idturno,Nombreturno in regturno:
                                  reservDisponibles.append((idsala,idturno))
                                  combinaciones_reservaciones_disponibles = set(reservDisponibles)
                          salas_turnos_disponibles = sorted(list(combinaciones_reservaciones_disponibles-reservaciones_ocupadas))
                          
                          print(combinaciones_reservaciones_disponibles)
                          print(reservaciones_ocupadas)
                          
                          print("\n Las opciones disponibles para rentar en esa fecha son : ")
                          print(f"Salas que se encuentran disponibles para rentar el {fechaBuscada}\n")
                          print("***Salas***\t\t***Turnos***")
                      if regsala:
                        for idsala,salas,idturno,Nombreturno in salas_turnos_disponibles:
                          print(f"{idsala}\t{salas}\t{idturno}\t{Nombreturno}")
                except Error as e:
                  print (e)
                except:
                  print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                finally:
                  conn.close()
            except Error as e:
              print (e)
            except:
              print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
              conn.close()
            break
          continue

        if (submenur.upper()=="D"):
          while True:
            try:
              with sqlite3.connect("PIA.db") as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM RESERVACION ORDER BY idreserv")
                regreserva = cursor.fetchall()
                if regreserva:
                    print("")
                    print("{:<20} {:<20} {:<20} {:<40} {:<20} {:<20}".format('idreserv','nombrereservacion','fechareservacion','idcliente','idturno','idsala' ))
                    print("*" * 150)
                    for idreserv ,nombrereservacion,fechareservacion,idcliente,idturno,idsala in regreserva:
                      print("{:<20} {:<20} {:<20} {:<40} {:<20} {:<20}".format(idreserv ,nombrereservacion,fechareservacion,idcliente,idturno,idsala))
                      continue
                    EliminarReg=int(input("Ingrese la clave del registro que desea eliminar : "))
                    valor={"Eliminado":EliminarReg}
                    cursor.execute("SELECT * FROM RESERVACION WHERE idreserv=:Eliminado",valor)
                    regreserva = cursor.fetchall()
                    if regreserva:
                      elire =input("Desea eliminar la reservacion [Si/No ]: ")
                      if (not elire.upper() in "SN"):
                          print("Opcion no valida")
                      if (elire.upper() == "S"):
                        try: #try para prevenir error de mal fecha ingresada
                          fecape=input("Ingrese la fecha a eliminar: \n")
                          if fecape=='':
                            print("No debe omitirse la fecha de la reservacion")
                            continue
                          else:
                            fechre=datetime.datetime.strptime(fecape, "%d/%m/%Y").date()  #fecha a eliminar
                        except ValueError:
                            print("Respuesta no valida: Solo se acepta fecha")
                        fechac=datetime.datetime.today().date()                                        #fecha actual
                        fechli= fechre + datetime.timedelta(days=-2)                         #fecha 3 dias antes de reservacion
                        if fechac >=fechli:                                                      #Compara que la fecha sea valida
                            print('Fecha no valida: Ingrese una fecha con tres dias de anticipacion.')
                        else:
                          cursor.execute("DELETE FROM reservacion WHERE idreserv = :Eliminado",valor)
                          print("Se ha eliminado la reservacion ")
                          break
                      elif (elire.upper() == "N"):
                        break
            except Error as e:
              print(e)
            except:
              print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
              conn.close()
            break
          continue

        if(submenur.upper()=="B"):
          while True:
            try:
              with sqlite3.connect("PIA.db") as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM RESERVACION ORDER BY idreserv")
                regreserva = cursor.fetchall()
                if regreserva:
                    print("")
                    print("{:<20} {:<20} {:<20} {:<40} {:<20} {:<20}".format('idreserv','nombrereservacion','fechareservacion','idcliente','idturno','idsala' ))
                    print("*" * 150)
                    for idreserv ,nombrereservacion,fechareservacion,idcliente,idturno,idsala in regreserva:
                      print("{:<20} {:<20} {:<20} {:<40} {:<20} {:<20}".format(idreserv ,nombrereservacion,fechareservacion,idcliente,idturno,idsala))
                      continue
                    buscarclave=int(input("Ingrese la clave del registro que va a editar : "))
                    valor={"Clavereser":buscarclave}
                    cursor.execute("SELECT * FROM RESERVACION WHERE idreserv= :Clavereser",valor)
                    regreserva = cursor.fetchall()
                    if regreserva:
                        print("")
                        nuevo_nombre=input("Ingrese el nuevo nombre de la reservacion  : ")
                        valores = {"Nombrereservacion":nuevo_nombre,"Clavereser":buscarclave}
                        cursor.execute("UPDATE RESERVACION SET nombrereservacion = :Nombrereservacion WHERE idreserv=:Clavereser",valores)
                        regreserva = cursor.fetchall()
                        print("El registro se edito exitosamente")
                        break
                    else:
                      if not idreserv in regreserva:
                        print("No existe este registro")
                        continue
            except Error as e:
              print (e)
            except:
              print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
              conn.close()
            break
          continue
        if (submenur.upper()=="A"):
          try: #try para prevenir error de mal fecha ingresada
              fecha_capturada=input("Ingrese la fecha para la reservacion: \n")
              if fecha_capturada=='':
                print("No debe omitirse la fecha de la reservacion")
                continue
              else:
                fechareservacion=datetime.datetime.strptime(fecha_capturada, "%d/%m/%Y").date()  #fecha de reservacion
          except ValueError:
              print("Respuesta no valida: Solo se acepta fecha")
          fechaactual=datetime.date.today()                                  #fecha actual
          fechalimite= fechareservacion.day-fechaactual.day                        #fecha 2 dias antes de reservacion
          if fechalimite<= 2:                                                      #Compara que la fecha sea valida
              print('Fecha no valida: Ingrese una fecha con dos dias de anticipacion.')
          else:
            NomRes=str(input("Ingrese el nombre de la reservacion : "))
            if NomRes=='':
              print("No debe omitirse el nombre de la reservacion")
              continue
            else:
              try:
                with sqlite3.connect("PIA.db") as conn:
                  cursor = conn.cursor()
                  cursor.execute("SELECT * FROM TURNO ORDER BY idturno")
                  regturno = cursor.fetchall()
                  #Procedemos a evaluar si hay registros en la respuesta
                  if regturno:
                    print("\nTURNOS....")
                    print("idturno\tNombreturno")
                    print("--" * 30)
                    for idturno, Nombreturno in regturno:
                      print(f"{idturno:^10}\t{Nombreturno}")
                    Turno=int(input("Ingrese el turno que desea reservar : "))
                    if Turno=='':
                      print("No debe omitirse el turno")
                      continue
                    else:
                      valores = {"Claveturno":Turno}
                      cursor.execute("SELECT * FROM TURNO WHERE idturno = :Claveturno", valores)
                      regturno = cursor.fetchall()
                      if regturno:
                        for Claveturno, Turno in regturno:
                          print("SE REGISTRO CORRECTAMENTE")
                          continue
                      else:
                        print("No se encontraron registros en la respuesta")
                        continue
              except Error as e:
                print(e)
              except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
              finally:
                conn.close()
              while True:
                try:
                  with sqlite3.connect("PIA.db") as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT * FROM  registrosala ORDER BY idsala")
                    regsala = cursor.fetchall()
                    if regsala:
                      print("\nSALAS REGISTRADAS....")
                      print("idsala\tsalas\tcupo_sala")
                      print("--" * 30)
                      for idsala, salas, cupo_sala in regsala:
                        print(f"{idsala:^10}\t{salas:^10}\t{cupo_sala}")
                      Sala=int(input("Ingresa el numero de la sala que desea reservar : "))
                      if Sala=='':
                        print("No debe omitirse el turno")
                        continue
                      else:
                        valores = {"Clavesala":Sala}
                        cursor.execute("SELECT * FROM registrosala WHERE idsala = :Clavesala", valores)
                        regsala = cursor.fetchall()
                        if regsala:
                          for Clavesala, salas, cupo_sala in regsala:
                            print("SE REGISTRO CORRECTAMENTE")
                            continue
                        else:
                          print("No se encontraron registros en la respuesta")
                          continue
                except Error as e:
                  print(e)
                except:
                  print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                finally:
                  conn.close()
                while True:
                  try:
                    with sqlite3.connect("PIA.db") as conn:
                      cursor = conn.cursor()
                      cursor.execute("SELECT * FROM   registrocliente ORDER BY idcliente")
                      regclie = cursor.fetchall()
                      if regclie:
                        print("\nCLIENTES REGISTRADOS....")
                        print("idcliente\tclientes")
                        print("--" * 30)
                        for idcliente, clientes in regclie:
                          print(f"{idcliente:^10}\t{clientes:^10}")
                        Cliente=int(input("Ingresa el id del cliente a reservar: "))
                        if Cliente=='':
                          print("No debe omitirse el cliente")
                          continue
                        else:
                          valores = {"Clavecliente":Cliente}
                          cursor.execute("SELECT * FROM registrocliente WHERE idcliente = :Clavecliente", valores)
                          regclie = cursor.fetchall()
                          if regclie:
                            for Clavecliente, clientes in regclie:
                              print("SE REGISTRO CORRECTAMENTE")
                              continue
                          else:
                            print("No se encontraron registros en la respuesta")
                            continue
                          try:
                              with sqlite3.connect("PIA.db") as conn:
                                cursor=conn.cursor()
                                valores={"Nombrereservacion":NomRes, "fechareservacion":fechareservacion, "idcliente":Cliente,"idturno":Turno,"idsala":Sala}
                                cursor.execute("INSERT INTO RESERVACION(nombrereservacion,fechareservacion,idcliente,idturno,idsala) VALUES(:Nombrereservacion,:fechareservacion,:idcliente,:idturno,:idsala)", valores)
                                print("Registro ha sido exitoso")
                          except Error as e:
                            print(e)
                          except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                          finally:
                            conn.close()
                  except Error as e:
                    print(e)
                  except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                  finally:
                    conn.close()
                  break
                break
              break
            break
          continue
        break
    if (opmenu.upper()=="D"):
      while True:
        print('*'*14)
        print('SUBMENU DE REPORTE')
        print('*'*14)
        print('A. REPORTE EN PANTALLA PARA RESERVACION DE UNA FECHA')
        print('B. EXPORTAR REPORTE EN EXCEL')
        print('C. REGRESAR AL MENU')
        submenure=input('\nEscoge la opción: ')
        if (not submenure.upper() in "ABC"):
            print("Opción no válida, intenta de nuevo.")
            continue
        if (submenure.upper()=="C"):
          break
          continue
        if (submenure.upper()=="A"):
          while True:
            try:
              fechaRep = input("Ingrese la fecha del evento (dd/mm/aaaa): ")
              if fechaRep =='':
                print("No debe omitirse la fecha de la reservacion")
                continue
              fecRer = datetime.datetime.strptime(fechaRep, "%d/%m/%Y").date()
              try:
                with sqlite3.connect("PIA.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                  cursor = conn.cursor()
                  valor = {"fecha":fecRer}
                  cursor.execute("SELECT nombrereservacion,DATE(fechareservacion),idcliente,idturno,idsala FROM RESERVACION WHERE DATE(fechareservacion) = :fecha;", valor)
                  regreserva = cursor.fetchall()
                  print("--"*48)
                  print("-" + " "*23 + f"REPORTE DE RESERVACIONES PARA EL DÍA {fecRer}" + " "*23 + "-")
                  print("--"*48)
                  print("{:<23} {:<23}  {:<23} {:<23} {:<23}".format('FECHA','SALA','CLIENTE','EVENTO', 'TURNO' ))
                  print("--"*48)
                  if regreserva:
                    for nombrereservacion,fechareservacion,idcliente,idturno,idsala in regreserva:
                      print(f"{fechareservacion}\t\t\t{idsala}\t\t\t{idcliente}\t\t\t{nombrereservacion}\t\t\t{idturno}")   
                      print("")
                      print("-"*40 + f"FIN DEL REPORTE PARA EL DÍA {fecRer}" + "-"*40)
                  else:
                    print(f"No hay eventos para este dia: {fecRer}")
              except sqlite3.Error as e:
                print (e)
              except Exception:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
              finally:
                conn.close()
            except ValueError:
              print('Respuesta no valida: solo se acepta fecha')
            break
          continue
        if (submenure.upper()=="B"):
          while True:
            try:
              fecExcel = input("Ingrese la fecha del evento (dd/mm/aaaa): ")
              if fecExcel =='':
                print("No debe omitirse la fecha de la reservacion")
                continue
              fecExR = datetime.datetime.strptime(fecExcel, "%d/%m/%Y").date()
              try:
                with sqlite3.connect("PIA.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                  cursor = conn.cursor()
                  valor = {"fecha":fecExR}
                  cursor.execute("SELECT nombrereservacion,DATE(fechareservacion),idcliente,idturno,idsala FROM RESERVACION WHERE DATE(fechareservacion)=:fecha", valor)
                  cursor.execute("SELECT salas FROM ")
                  regreserva = cursor.fetchall()
                  if regreserva:
                    valores_Reserva = [(fechareservacion,idsala,idcliente,nombrereservacion,idturno)]
                    hoja["B1"].value = "Hoy es:"
                    hoja["C1"].value = datetime.datetime.now().strftime("%d/%m/%Y")
                    hoja.merge_cells('A2:K2')
                    hoja["A2"].value=("--"*100)
                    hoja.merge_cells('A3:B3')
                    hoja["A3"].value=("--"*100)
                    hoja.merge_cells('C3:G3')
                    hoja["C3"].value="REPORTE DE RESERVACIONES PARA EL DIA "
                    hoja.merge_cells('H3:I3')
                    hoja["H3"].value=f"{fecExR}"
                    hoja.merge_cells('J3:K3')
                    hoja["J3"].value=("--"*100)
                    hoja["A3"].value="FECHA"
                    hoja["B5"].value="SALA"
                    hoja["C5"].value="CLIENTE"
                    hoja["D5"].value="EVENTO"
                    hoja["E5"].value="TURNO"
                    for nombrereservacion,fechareservacion,idcliente,idturno,idsala in regreserva:
                      valores_Reserva = [(fechareservacion,idsala,idcliente,nombrereservacion,idturno)]
                      for valor in valores_Reserva:
                        hoja.append(valor)
                      libro.save(f"ReporteDeReservaciones{fecExR}.xlsx")
                      print("\nLibro del reporte creado exitosamente!")
                  else:
                    print(f"No hay eventos para: {fecExR}")
                    continue
              except sqlite3.Error as e:
                print (e)
              except Exception:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
              finally:
                conn.close()
            except ValueError:
              print('Respuesta no valida: solo se acepta fecha')
            break
          continue
